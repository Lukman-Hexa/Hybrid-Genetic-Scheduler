from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def export_to_excel(
    final_schedule, structured_result, filepath, semester_type="ganjil"
):
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("Summary", 0)
    setup_summary_sheet(summary_ws, final_schedule, structured_result, semester_type)

    semesters = sorted(structured_result.keys())
    for idx, semester in enumerate(semesters):
        ws_name = f"Semester {semester}"
        ws = wb.create_sheet(ws_name, idx + 1)
        setup_semester_sheet(ws, structured_result[semester], semester)

    wb.save(filepath)
    print(f"✓ File Excel disimpan: {filepath}")


def setup_summary_sheet(ws, final_schedule, structured_result, semester_type):
    ws["A1"] = f"JADWAL KULIAH - RINGKASAN ({semester_type.upper()})"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:F1")

    row = 3
    ws[f"A{row}"] = "Tipe Semester:"
    ws[f"B{row}"] = semester_type.upper()
    ws[f"B{row}"].font = Font(bold=True, color="0000FF")

    row += 1
    ws[f"A{row}"] = "Total Sesi:"
    ws[f"B{row}"] = len(final_schedule)

    row += 1
    ws[f"A{row}"] = "Total Semester:"
    ws[f"B{row}"] = len(structured_result)

    row += 3
    ws[f"A{row}"] = "Distribusi Per Semester:"
    ws[f"A{row}"].font = Font(bold=True)

    row += 1
    for sem in sorted(structured_result.keys()):
        total_sessions = sum(
            len(sessions) for sessions in structured_result[sem].values()
        )
        ws[f"A{row}"] = f"Semester {sem}:"
        ws[f"B{row}"] = total_sessions
        ws[f"B{row}"].font = Font(bold=True)
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15


def setup_semester_sheet(ws, semester_data, semester_num):
    ws["A1"] = f"JADWAL SEMESTER {semester_num}"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    ws.merge_cells("A1:G1")

    headers = [
        "Hari",
        "Jam Mulai",
        "Jam Selesai",
        "Mata Kuliah",
        "Kelas",
        "Dosen",
        "Ruangan",
    ]
    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    header_font = Font(bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 4
    day_order = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat"]

    for day in day_order:
        if day in semester_data:
            for session in semester_data[day]:
                ws.cell(row=row, column=1).value = day
                ws.cell(row=row, column=2).value = session["start_time"]
                ws.cell(row=row, column=3).value = session["end_time"]
                ws.cell(row=row, column=4).value = session["course"]
                ws.cell(row=row, column=5).value = session["kelas"]
                ws.cell(row=row, column=6).value = session["dosen"]
                ws.cell(row=row, column=7).value = session["room"]

                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                row += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 10


def export_to_excel_buffer(
    final_schedule, structured_result, buffer, semester_type="ganjil"
):
    """Export jadwal ke BytesIO buffer — tanpa menyimpan file ke disk"""
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("Summary", 0)
    setup_summary_sheet(summary_ws, final_schedule, structured_result, semester_type)

    semesters = sorted(structured_result.keys())
    for idx, semester in enumerate(semesters):
        ws_name = f"Semester {semester}"
        ws = wb.create_sheet(ws_name, idx + 1)
        setup_semester_sheet(ws, structured_result[semester], semester)

    wb.save(buffer)
